#!/usr/bin/env python3
"""Build the privacy-safe data payload for the Japanese alignment website."""

from __future__ import annotations

import json
import math
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from scipy.stats import pearsonr, spearmanr, ttest_rel, wilcoxon


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "research-insights-7f3c9a" / "gesture_ratings_all_data_and_analysis.xlsx"
OUTPUTS = [
    ROOT / "japanese-alignment-data.json",
    ROOT / "japanese-alignment-data.js",
]

DIMENSIONS = [
    ("iconicity", "Iconicity"),
    ("sensorimotor_imagery", "Sensorimotor imagery"),
    ("motional_salience_gesture", "Motional salience"),
    ("emotional_salience_facial_expression", "Facial emotion"),
    ("gesture_complexity_fit", "Complexity fit"),
    ("cultural_familiarity", "Cultural familiarity"),
    ("enactment_potential", "Enactment potential"),
]

CONDITIONS = {
    "baseline_flash": {
        "label": "Baseline Flash-Lite",
        "model": "Gemini 3.1 Flash-Lite Preview",
        "directory": ROOT / "results" / "all_rating_flash",
        "family": "flash",
        "condition": "baseline",
    },
    "japanese_flash_lite": {
        "label": "Japanese-perspective Flash-Lite",
        "model": "Gemini 3.1 Flash-Lite Preview",
        "directory": ROOT / "results" / "japanese_perspective_flash_lite",
        "family": "flash",
        "condition": "japanese",
    },
    "baseline_pro": {
        "label": "Baseline Pro",
        "model": "Gemini 3.1 Pro Preview",
        "directory": ROOT / "results" / "all_rating_pro",
        "family": "pro",
        "condition": "baseline",
    },
    "japanese_pro": {
        "label": "Japanese-perspective Pro",
        "model": "Gemini 3.1 Pro Preview",
        "directory": ROOT / "results" / "japanese_perspective_pro",
        "family": "pro",
        "condition": "japanese",
    },
}


def human_means() -> tuple[dict[str, dict], dict]:
    sheet = load_workbook(WORKBOOK, read_only=True, data_only=True)["Human Ratings"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    values: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    raters: set[str] = set()
    response_count = 0

    for row in rows:
        record = dict(zip(headers, row))
        if record["language"] != "ja":
            continue
        response_count += 1
        raters.add(record["rater_code"])
        for key, _ in DIMENSIONS:
            values[record["video_title"]][key].append(float(record[key]))

    result = {}
    for title, dimensions in values.items():
        result[title] = {
            "n": len(next(iter(dimensions.values()))),
            "means": {
                key: sum(scores) / len(scores)
                for key, scores in dimensions.items()
            },
        }
    return result, {
        "videos": len(result),
        "raters": len(raters),
        "video_responses": response_count,
        "individual_scores": response_count * len(DIMENSIONS),
        "ratings_per_video_min": min(item["n"] for item in result.values()),
        "ratings_per_video_median": sorted(item["n"] for item in result.values())[len(result) // 2],
        "ratings_per_video_max": max(item["n"] for item in result.values()),
    }


def model_scores(directory: Path, titles: list[str]) -> dict[str, dict[str, float]]:
    result = {}
    for title in titles:
        path = directory / f"{Path(title).stem}.rating.json"
        payload = json.loads(path.read_text(encoding="utf-8"))
        result[title] = {
            key: float(payload["ratings"][key]["score"])
            for key, _ in DIMENSIONS
        }
    return result


def stats(pairs: list[tuple[float, float]]) -> dict:
    human = [pair[0] for pair in pairs]
    model = [pair[1] for pair in pairs]
    differences = [model_value - human_value for human_value, model_value in pairs]
    return {
        "n": len(pairs),
        "mae": sum(abs(value) for value in differences) / len(differences),
        "rmse": math.sqrt(sum(value * value for value in differences) / len(differences)),
        "bias": sum(differences) / len(differences),
        "pearson_r": float(pearsonr(human, model).statistic),
        "spearman_rho": float(spearmanr(human, model).statistic),
    }


def main() -> int:
    human, metadata = human_means()
    titles = sorted(human)
    model_data = {
        key: model_scores(config["directory"], titles)
        for key, config in CONDITIONS.items()
    }

    condition_stats = {}
    for condition, config in CONDITIONS.items():
        dimension_stats = {}
        all_pairs = []
        for key, label in DIMENSIONS:
            pairs = [
                (human[title]["means"][key], model_data[condition][title][key])
                for title in titles
            ]
            dimension_stats[key] = {"label": label, **stats(pairs)}
            all_pairs.extend(pairs)
        condition_stats[condition] = {
            "label": config["label"],
            "model": config["model"],
            "family": config["family"],
            "condition": config["condition"],
            "overall": stats(all_pairs),
            "dimensions": dimension_stats,
            "mean_dimension_pearson_r": sum(
                item["pearson_r"] for item in dimension_stats.values()
            )
            / len(DIMENSIONS),
            "mean_dimension_spearman_rho": sum(
                item["spearman_rho"] for item in dimension_stats.values()
            )
            / len(DIMENSIONS),
        }

    comparisons = {}
    for family, baseline_key, japanese_key in [
        ("pro", "baseline_pro", "japanese_pro"),
        ("flash", "baseline_flash", "japanese_flash_lite"),
    ]:
        baseline_video_mae = []
        japanese_video_mae = []
        for title in titles:
            baseline_video_mae.append(
                sum(
                    abs(model_data[baseline_key][title][key] - human[title]["means"][key])
                    for key, _ in DIMENSIONS
                )
                / len(DIMENSIONS)
            )
            japanese_video_mae.append(
                sum(
                    abs(model_data[japanese_key][title][key] - human[title]["means"][key])
                    for key, _ in DIMENSIONS
                )
                / len(DIMENSIONS)
            )
        differences = [
            japanese - baseline
            for baseline, japanese in zip(baseline_video_mae, japanese_video_mae)
        ]
        paired_t = ttest_rel(japanese_video_mae, baseline_video_mae)
        signed_rank = wilcoxon(differences)
        comparisons[family] = {
            "baseline": baseline_key,
            "japanese": japanese_key,
            "mean_video_mae_change": sum(differences) / len(differences),
            "improved_videos": sum(value < 0 for value in differences),
            "unchanged_videos": sum(value == 0 for value in differences),
            "worsened_videos": sum(value > 0 for value in differences),
            "paired_t_p": float(paired_t.pvalue),
            "wilcoxon_p": float(signed_rank.pvalue),
        }

    video_points = {}
    for key, label in DIMENSIONS:
        video_points[key] = {
            "label": label,
            "points": [
                {
                    "title": title,
                    "n": human[title]["n"],
                    "human": human[title]["means"][key],
                    **{
                        condition: model_data[condition][title][key]
                        for condition in CONDITIONS
                    },
                }
                for title in titles
            ],
        }

    payload = {
        "metadata": {
            **metadata,
            "dimensions": len(DIMENSIONS),
            "analysis_date": "2026-07-27",
            "benchmark": "Japanese-language human responses only",
            "scale": "1–5",
        },
        "conditions": condition_stats,
        "comparisons": comparisons,
        "video_points": video_points,
        "notes": {
            "primary_comparison": (
                "Baseline Pro versus Japanese-perspective Pro. Both use the Pro model "
                "condition; only the rating instructions and response language change."
            ),
            "flash_caveat": (
                "The Flash-Lite comparison uses Gemini 3.1 Flash-Lite Preview in both "
                "conditions; the rating instructions and response language change."
            ),
            "correlation_caveat": (
                "Overall correlations pool all seven dimensions. Dimension-specific "
                "correlations are required to assess within-dimension ranking."
            ),
        },
    }

    serialized = json.dumps(payload, indent=2, ensure_ascii=False) + "\n"
    for output in OUTPUTS:
        output.parent.mkdir(parents=True, exist_ok=True)
        if output.suffix == ".js":
            output.write_text(
                f"window.JAPANESE_ALIGNMENT_DATA = {serialized.rstrip()};\n",
                encoding="utf-8",
            )
        else:
            output.write_text(serialized, encoding="utf-8")
        print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
